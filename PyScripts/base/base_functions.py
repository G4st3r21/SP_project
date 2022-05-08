from openpyxl import *
import psycopg2
from os import system

from openpyxl.utils import range_boundaries


def db_conn(db_name='project-cs', table_name='other'):
    print(f"Подключение к базе данных '{db_name}'. Таблица - '{table_name}'")
    conn = psycopg2.connect(dbname=db_name, user='cs-main',
                            password='eid4Uepo', host='cs-vm-postgre.cs.vsu.ru')
    cursor = conn.cursor()

    return cursor, conn


def clear_console():
    system('clear')


def xlsx_connect(path):
    wb = load_workbook(path)

    return wb


def parser_init(file_name, sheet_number, first_str_number):
    cur, conn = db_conn(table_name="All Tables")
    wb = xlsx_connect(file_name)
    ws = wb.worksheets[sheet_number - 1]
    unmerge_all_cells(ws)

    columns = list(ws.columns)
    rows = list(ws.rows)

    return columns, rows[first_str_number - 1:], cur, conn


def read_sheet_number():
    return int(input("Номер листа: ")) - 1


def read_first_str_number():
    return int(input("Номер строки для начала парсинга: ")) - 1


def unmerge_all_cells(ws):
    try:
        while len(ws.merged_cells.ranges) > 0:
            min_col, min_row, max_col, max_row = range_boundaries(str(ws.merged_cells.ranges[0]))
            top_left_cell_value = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(str(ws.merged_cells.ranges[0]))
            for row in ws.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = top_left_cell_value
    except Exception:
        pass

def partition(string):
    parts = string.split()
    str_list = list()
    k = 0
    for i in range(len(parts) - 1):
        if ((parts[i][0] in 'АБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЭЮЯ' and parts[i][1] != '.') or
            (len(parts[i]) >= 5 and parts[i][0] in 'АБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЭЮЯ' and parts[i][1] == '.' and parts[i][2] in 'АБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЭЮЯ' and parts[i][3] == '.' and parts[i][4] in 'АБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЭЮЯ')) \
                and (',' in parts[i] or '.' in parts[i] or parts[i + 1][0] in 'АБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЭЮЯ' and parts[i + 1][1] != '.'):
            str_list.append(' '.join(parts[k:i + 1]))
            k = i + 1
    str_list.append(' '.join(parts[k:]))
    for j in range(len(str_list)):
        if not (str_list[j] in 'АБВГДЕЁЖЗИЙКЛМНОПРСТУФХЦЧШЩЭЮЯ'):
            str_list[j] = str_list[j].replace(str_list[j].split()[0], str_list[j].split()[0].capitalize())
        if str_list[j][-1] in '.,;:-+=?!' or str_list[j].endswith('\n'):
            str_list[j] = str_list[j][:-1]
        if str_list[j].split()[0] == 'Врио' or str_list[j].split()[0] == 'врио':
            str_list[j] = str_list[j].replace(str_list[j].split()[0], 'ВРИО')

    return str_list

def parts(string, sym):
    parts = string.split(sym)
    for i in range(len(parts)):
        parts[i] = format_title(parts[i]).strip()
        if parts[i].split()[0] == 'Врио' or parts[i].split()[0] == 'врио':
            parts[i] = parts[i].replace(parts[i].split()[0], 'ВРИО')

    return parts

def rreplace(s, old, new, occurrence):
    li = s.rsplit(old, occurrence)
    return new.join(li)

def format_title(string):
    if '\n' in string and not string.endswith('\n'):
        string = string.replace('\n', ' ')
    elif string[-1] == '\n':
        string = string.replace('\n', '')

    string = string.strip()
    if string[-1] == '.':
        string = rreplace(string, '.', '', 1)

    string = string.replace(string.split()[0], string.split()[0].capitalize())
    # string.replace(string[0], string[0].upper())
    # string.capitalize()
    return ' '.join(string.split())

