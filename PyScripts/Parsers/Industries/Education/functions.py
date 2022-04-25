import psycopg2
from openpyxl.utils import range_boundaries


# !!!!!!!!!
def db_conn():
    conn = psycopg2.connect(dbname='project-cs', user='cs-main',
                            password='eid4Uepo', host='cs-vm-postgre.cs.vsu.ru')
    cursor = conn.cursor()

    return cursor, conn


def read_sheet_number():
    return int(input("Номер листа: ")) - 1


def read_first_str_number():
    return int(input("Номер строки для начала парсинга: ")) - 1

expense_items = [
    "Всего",
    "Государственные капитальные вложения, всего",
    "Государственные капитальные вложения (объекты капитального строительства и недвижимое имущество)",
    "субсидии местным бюджетам на софинансирование объектов муниципальной собственности",
    "НИОКР",
    "Прочие расходы",
    "бюджетные инвестиции на финансирование объектов областной собственности",
    "Государственные капитальные вложения (за исключением объектов капитального строительства и объектов недвижимого имущества)",
    "Государственные капитальные вложения"
]

categories = [
    "согласно закону Воронежской области об областном бюджете на отчетную дату текущего года",
    "согласно бюджетной росписи расходов областного бюджета на отчетную дату текущего года",
    "доведенный департаментом финансов Воронежской области предельный объем финансирования "
    "(поквартальный кассовый план на отчетную дату нарастающим итогом)",
    "кассовое исполнение (на отчетную дату нарастающим итогом)",
    "поквартальный кассовый план на отчетную дату нарастающим итогом",
    "Уровень освоения бюджетных ассигнований"
]

def roundOrNone(value):
    if value is not None:
        return round(value, 2)
    else:
        return None

def unmerge_all_cells(wb):
    while len(wb.merged_cells.ranges) > 0:
        print(len(wb.merged_cells.ranges))
        min_col, min_row, max_col, max_row = range_boundaries(str(wb.merged_cells.ranges[0]))
        top_left_cell_value = wb.ws.cell(row=min_row, column=min_col).value
        wb.ws.unmerge_cells(str(wb.merged_cells.ranges[0]))
    for row in wb.ws.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
        for cell in row:
            cell.value = top_left_cell_value

kbk = []
response = []

def find_response():
    print()
