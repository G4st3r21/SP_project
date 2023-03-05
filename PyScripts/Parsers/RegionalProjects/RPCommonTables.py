from base_functions_RP import get_file_names, get_code, format_title, format_date

from openpyxl import load_workbook
from PyScripts.base.base_functions import db_conn
from PyScripts.TableClasses.SPTables.SPTable import SPTable
from PyScripts.TableClasses.SPTables.SPTableArbitrary import SPTableArbitrary

cur, conn = db_conn(table_name="All Tables")

FedProject = SPTable('fed_project', cur, conn)
RegProject = SPTableArbitrary('reg_project', cur, conn)
AdditionInRegProject = SPTableArbitrary('addition_in_reg_project', cur, conn)
RegProject2019 = SPTableArbitrary('reg_project2019', cur, conn)
RegProjectsAndGosprogram2019 = SPTableArbitrary('reg_project_and_gosprogram2019', cur, conn)
RegResponseFio2019 = SPTableArbitrary('reg_response_fio2019', cur, conn)
RegProjectsAndResponses2019 = SPTableArbitrary('reg_projects_and_responses2019', cur, conn)

ResponseObj = SPTable('response_obj', cur, conn)


def get_response_obj_id(obj):
    """Returns id of a response object if object exists or adds it and returns id. By i.korn."""
    title_response_obj = ""
    if "департамент" in obj:
        str = obj[obj.find("департамент"):].split(' ')
        str[0] = str[0][:len("департамент")]
        str[0] = str[0][0].upper() + str[0][1:len(str[0])]
        for word in str:
            title_response_obj += (word + " ")
    if "правительств" in obj:
        str = obj[obj.find("правительств"):].split(' ')
        str[0] = str[0][:len("правительств")] + "о"
        str[0] = str[0][0].upper() + str[0][1:len(str[0])]
        for word in str:
            title_response_obj += (word + " ")
    response_obj_id = ResponseObj.add(title_response_obj.strip())
    ResponseObj.commit()
    if response_obj_id:
        return response_obj_id
    else:
        return None


def get_position(obj):
    """Returns position without object. By i.korn."""
    position = ""
    if "департамент" in obj:
        position = obj[:obj.find("департамент")].strip()
    if "правительств" in obj:
        position = obj[:obj.find("правительств")].strip()
    return position


def fill_fed_project(sheet):
    """Temporary function to fill only FED_PROJECT tables"""
    for row in sheet.iter_rows(min_row=6, max_row=9):
        if row[0].value is not None:
            if 'Наименование федерального проекта' in format_title(row[0].value):
                for j in range(1, len(row)):
                    if row[j].value is not None:
                        title_federal_project = format_title(row[j].value)
                        FedProject.add(title_federal_project)
                        FedProject.commit()


def parse_schemes():
    path = "D:/КСП/Tables/5_Regionalnye_proekty/5_Региональные проекты/Сравнительная таблица региональных проектов.xlsx"
    wb = load_workbook(path)
    sheet = wb.active
    schemes_dict = {}
    scheme_code = ""
    for row in sheet.iter_rows():
        if row[0].value is not None:
            if "Код" in row[0].value:
                scheme_code = row[1].value
            if "Схема" in row[0].value and scheme_code != "":
                schemes_dict[scheme_code] = row[1].value
    return schemes_dict


def fill_reg_project(sheet, code, schema):
    """Fills REG_PROJECT and FED_PROJECT tables"""
    title = format_title(sheet['A6'].value)
    for row in sheet.iter_rows(max_row=14):
        if row[0].value is not None:
            if 'Наименование федерального проекта' in format_title(row[0].value):
                for j in range(1, len(row)):
                    if row[j].value is not None:
                        title_federal_project = format_title(row[j].value)
                        id_fed_project = FedProject.add(title_federal_project)
                        FedProject.commit()
            if 'Краткое наименование регионального проекта' in format_title(row[0].value):
                for j in range(1, len(row)):
                    if row[j].value is not None:
                        if 'Срок начала и окончания проекта' in format_title(row[j].value):
                            for k in range(j + 1, len(row)):
                                if row[k].value is not None:
                                    dates = format_date(row[k].value)
                                    start_date, end_date = dates[0], dates[1]
                            break
                        else:
                            print(row[j].value)
                            short_title = row[j].value

    RegProject.add(code, id_fed_project, title, short_title, start_date, end_date, schema)
    RegProject.commit()


def fill_addition_in_reg_project(sheet, code_project):
    """Fills ADDITION_IN_REG_PROJECT table"""
    additional_info_title = False
    addition = ""
    for row in sheet.iter_rows():
        if row[0].value is not None:
            if "6. Дополнительная информация" in format_title(row[0].value):
                additional_info_title = True
                continue
            if "ПЛАН" in format_title(row[0].value) \
                    or "МЕТОДИКА" in format_title(row[0].value):
                additional_info_title = False
                break
            if additional_info_title:
                to_add = format_title(row[0].value)
                if not to_add.isnumeric():
                    if addition != "":
                        addition += " " + to_add
                    else:
                        addition += to_add
    if addition != "":
        AdditionInRegProject.add(code_project, addition)
        AdditionInRegProject.commit()


def fill_reg_response_fio2019(sheet):
    """Fills REG_RESPONSE_FIO2019 table. By i.korn."""
    # todo: people with no response_obj: F2, F3, G5, GA, P1, P5, T1, T2
    for row in sheet.iter_rows(min_row=8, max_row=13):
        if "Куратор регионального проекта" in format_title(row[0].value) \
                or "Руководитель регионального проекта" in format_title(row[0].value) \
                or "Администратор регионального проекта" in format_title(row[0].value):
            for j in range(1, len(row)):
                if row[j].value is not None:
                    response_obj_id = get_response_obj_id(row[j].value.split(", ")[1])
                    reg_response_fio = row[j].value.split(", ")[0]
                    position = get_position(row[j].value.split(", ")[1])
                    if response_obj_id and position:
                        RegResponseFio2019.add(response_obj_id, reg_response_fio, position)
                        RegResponseFio2019.commit()
                    break


def fill_reg_project2019(sheet, code):
    """Fills REG_PROJECT2019 table"""
    curator_id, manager_id, administrator_id = 0, 0, 0
    for row in sheet.iter_rows(max_row=14):
        if row[0].value is not None:
            if "Куратор регионального проекта" in format_title(row[0].value):
                for j in range(1, len(row)):
                    if row[j].value is not None:
                        fio = row[j].value.split(", ")[0]
                        args_dict = {'reg_response_fio': fio}
                        curator_tuple = RegResponseFio2019.get_by_custom_filter_expression(args_dict)
                        if curator_tuple:
                            curator_id = curator_tuple[0][0]
            if "Руководитель регионального проекта" in format_title(row[0].value):
                for j in range(1, len(row)):
                    if row[j].value is not None:
                        fio = row[j].value.split(", ")[0]
                        args_dict = {'reg_response_fio': fio}
                        manager_tuple = RegResponseFio2019.get_by_custom_filter_expression(args_dict)
                        if manager_tuple:
                            manager_id = manager_tuple[0][0]
            if "Администратор регионального проекта" in format_title(row[0].value):
                for j in range(1, len(row)):
                    if row[j].value is not None:
                        fio = row[j].value.split(", ")[0]
                        args_dict = {'reg_response_fio': fio}
                        administrator_tuple = RegResponseFio2019.get_by_custom_filter_expression(args_dict)
                        if administrator_tuple:
                            administrator_id = administrator_tuple[0][0]
    if curator_id and manager_id and administrator_id:
        RegProject2019.add(code, curator_id, manager_id, administrator_id)  # no serial, has_id=True in .add
        RegProject2019.commit()


def fill_reg_projects_and_responses2019(sheet):
    """Fills REG_PROJECTS_AND_RESPONSES2019 table"""
    for row in sheet.iter_rows():
        if "Участники регионального проекта" in row[0].value:
            while row[0].value is not None:
                code_reg_project = 1
                id_response_fio = 1
                role_in_reg_project = ""
                direct_supervisor_id = 1
                employment_percents = 0.50
                RegProjectsAndResponses2019.add(code_reg_project, id_response_fio, role_in_reg_project,
                                                direct_supervisor_id, employment_percents)
                RegProjectsAndResponses2019.commit()


def fill_reg_project_and_gosprogram2019(sheet):
    """Fills REG_PROJECT_AND_GOSPROGRAM2019 table"""


# args_dict = {'reg_response_fio': "Попов Владимир Борисович"}
# curator_tuple = RegResponseFio2019.get_by_custom_filter_expression(args_dict)
# curator_id = curator_tuple[0][0]
#
# print(type(curator_id))
path = "D:/КСП/Tables/5_Regionalnye_proekty/5_Региональные проекты/2019 паспорта по состоянию на 23.12.2019"
files = get_file_names(path)

for f in files:
    code = get_code(f)
    if code:
        full_path = path + "/" + f
        wb = load_workbook(full_path)

        # fill_reg_response_fio2019(wb.active)
        fill_reg_project2019(wb.active, code)
        print(code, "file is loaded to database")
