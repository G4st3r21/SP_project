import regex as regex

from base_functions_RP import get_file_names, get_code, format_title, format_date_single

from openpyxl import load_workbook
from PyScripts.base.base_functions import db_conn
from PyScripts.TableClasses.SPTables.SPTable import SPTable
from PyScripts.TableClasses.SPTables.SPTableArbitrary import SPTableArbitrary

cur, conn = db_conn(table_name="All Tables")
RegProject = SPTableArbitrary('reg_project', cur, conn)
RegPassportAims = SPTableArbitrary('reg_passport_aims', cur, conn)
FedIndicators2019 = SPTableArbitrary('fed_indicators2019', cur, conn)
TypesOfIndicators = SPTable('types_of_indicators', cur, conn)
RegAimsIndicatorsNames2019 = SPTableArbitrary('reg_aims_indicators_names2019', cur, conn)


def fill_reg_passport_aims(sheet, code_reg_project):
    """Fills reg_passport_aims table"""
    RegPassportAims = SPTableArbitrary('reg_passport_aims', cur, conn)
    passport_aim = ""
    aim_and_indicators = False
    for row in sheet.iter_rows(max_row=22):
        if row[0].value is not None:
            if "2. Цель и показатели регионального проекта" in format_title(row[0].value):
                aim_and_indicators = True
                continue
            if aim_and_indicators:
                passport_aim = format_title(row[0].value)
                break
    if passport_aim:
        RegPassportAims.add(code_reg_project, passport_aim)  # serial
        RegPassportAims.commit()


def fill_fed_indicators2019(sheet, code):
    """Fills fed_indicators2019 table"""
    args_dict = {'code': code}
    id_fed_project = RegProject.get_by_custom_filter_expression(args_dict)[0][1]
    title_indicators = []
    skip_aim = False
    fed_indicators = False
    for row in sheet.iter_rows():
        if row[0].value is not None:
            row_data = format_title(row[0].value)
            if "2. Цель и показатели регионального проекта" in row_data:
                skip_aim = True
                continue
            if skip_aim:
                skip_aim = False
                fed_indicators = True
                continue
            if "3. Результаты регионального проекта" in row_data:
                break
            if fed_indicators and row_data[0].isalpha() \
                    and "Отсутствует показатель федерального проекта" not in row_data:
                title_indicators.append(row_data)
    if title_indicators:
        # print(code, title_indicators)
        FedIndicators2019.add(id_fed_project, title_indicators[0])  # serial
        FedIndicators2019.commit()


def fill_reg_aims_indicators_names2019(sheet, code):
    """Fills reg_aims_indicators_names2019 table"""
    # serial id
    args_dict = {'code_reg_project': code}
    # can be found by code
    id_reg_aim = RegPassportAims.get_by_custom_filter_expression(args_dict)[0][0]

    fed_indicator_id = None  # FK to fed_indicators2019. can be found by title
    type_id = None  # FK to types_of_indicators

    title_indicator = None  # without units
    unit_of_measurement = None

    dynamics_id = None  # FK to dynamics
    base_value = None
    base_date = None

    skip_aim = False
    fed_indicators = False
    indicators = False

    TITLE_INDICATOR_INDEX = 0
    TYPE_INDICATOR_INDEX = 1
    BASE_VALUE_INDEX = 2
    BASE_DATE_INDEX = 3
    VALUE2019_INDEX = 4
    VALUE2020_INDEX = 5
    VALUE2021_INDEX = 6
    VALUE2022_INDEX = 7
    VALUE2023_INDEX = 8
    VALUE2024_INDEX = 9
    for row in sheet.iter_rows(min_row=14):
        if row[0].value is not None:
            row_data = format_title(row[0].value)
            if "2. Цель и показатели регионального проекта" in row_data:
                skip_aim = True
            elif "3. Результаты регионального проекта" in row_data:
                break
            elif skip_aim:
                skip_aim = False
                fed_indicators = True
            elif fed_indicators and row_data[0].isalpha():
                indicators = True
                if "Отсутствует показатель федерального проекта" not in row_data:
                    args_dict = {'title_indicator': row_data}
                    fed_indicator_id = FedIndicators2019.get_by_custom_filter_expression(args_dict)[0][0]
                else:
                    fed_indicator_id = None
            elif indicators and row_data[0].isdigit() and '.' in row_data:  # and not row_data.isdigit()
                index = 0
                for j in range(1, len(row)):
                    if row[j].value is not None and index < 5:
                        if index == TITLE_INDICATOR_INDEX:
                            row_data_in = format_title(row[j].value)
                            # done: process title
                            title_indicator = row_data_in

                            parce_units = regex.search(r"(?r), (\w+|\s+)+$", row_data_in)
                            if parce_units is not None:
                                parce_units = parce_units.group()
                                unit_of_measurement = parce_units[2:]
                                title_indicator = title_indicator.replace(parce_units, '')
                            else:
                                unit_of_measurement = None

                            if "(накопительным итогом)" in row_data_in:
                                dynamics_id = 1
                                title_indicator = title_indicator.replace('(накопительным итогом)', '')
                            elif "(нарастающим итогом)" in row_data_in:
                                dynamics_id = 1
                                title_indicator = title_indicator.replace('(нарастающим итогом)', '')
                            else:
                                dynamics_id = None
                        if index == TYPE_INDICATOR_INDEX:
                            row_data_in = format_title(row[j].value)
                            args_dict = {'type_of_indicator': row_data_in}
                            type_id = TypesOfIndicators.get_by_custom_filter_expression(args_dict)[0][0]
                        if index == BASE_VALUE_INDEX:
                            # done: cast base value
                            base_value = row[j].value
                        if index == BASE_DATE_INDEX:
                            row_data_in = format_title(row[j].value)
                            # done: cast date, alter format_date func for one date
                            base_date = format_date_single(row_data_in)  # format_date(row_data_in)
                            break
                        index += 1
                    elif row[j].value is not None and index < 11:
                        pass_values_dict = {}
                        if index == VALUE2019_INDEX:
                            pass_values_dict['4'] = row[j].value
                        elif index == VALUE2020_INDEX:
                            pass_values_dict['5'] = row[j].value
                        elif index == VALUE2021_INDEX:
                            pass_values_dict['6'] = row[j].value
                        elif index == VALUE2022_INDEX:
                            pass_values_dict['7'] = row[j].value
                        elif index == VALUE2023_INDEX:
                            pass_values_dict['8'] = row[j].value
                        elif index == VALUE2024_INDEX:
                            pass_values_dict['9'] = row[j].value
                        index += 1

                # todo: note that ER diagram has different order of arguments than the actual DB table has
                reg_indicator_id = RegAimsIndicatorsNames2019.add(title_indicator, unit_of_measurement, id_reg_aim,
                                                                  type_id, base_value, base_date,
                                                                  dynamics_id, fed_indicator_id)
                RegAimsIndicatorsNames2019.commit()
                print(reg_indicator_id)
                # todo: if add returns reg_indicator_id here we can also fill reg_indicators_passport_values in schema


def fill_types_of_indicators():
    """Fills types_of_indicators table"""
    TypesOfIndicators = SPTable('types_of_indicators', cur, conn)
    # todo: find out if there's another types of indicators
    TypesOfIndicators.add("Дополнительный показатель")
    TypesOfIndicators.commit()


# fill_types_of_indicators()


def fill_dynamics():
    """Fills dynamics table"""
    # id serial
    dynamics = ""


def fill_indicators_comments2019():
    """Fills indicators_comments2019 table for each individual schema"""
    id_reg_indicator = 0
    comment = ""


def fill_indicators_additional_comments2019():
    """Fills indicators_additional_comments2019 table for each individual schema"""
    id_reg_indicator = 0
    additional_comment = ""


def fill_reg_aims_indicators2019():
    """Fills reg_aims_indicators2019 table for each individual schema"""
    # serial id
    indicator_id = 0
    is_deviation = 0
    plan_fact = False
    previous_this_year = False
    quarter_number = 0
    indicator = 0.0


def fills_reg_indicators_passport_values2019():
    """Fills reg_indicators_passport_values2019 for each individual schema"""
    # serial id
    reg_indicator_id = 0
    year_id = 0
    value = 0.0


path = "D:/Библиотеки/КСП/Tables/5_Regionalnye_proekty/5_Региональные проекты/2019 паспорта по состоянию на 23.12.2019"
files = get_file_names(path)

for f in files:
    code = get_code(f)
    if code == 'E2':
        full_path = path + "/" + f
        wb = load_workbook(full_path)
        #args_dict = {'title_indicator': "Доля детей в возрасте от 5 до 18 лет, охваченных дополнительным образованием"}
        #fed_indicator_id = FedIndicators2019.get_by_custom_filter_expression(args_dict)[0][0]
        fill_reg_aims_indicators_names2019(wb.active, code)
        #print(fed_indicator_id)
        print(code, "file is loaded to database")
