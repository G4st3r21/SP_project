from openpyxl import load_workbook
import psycopg2

table_names = ["CUR", "TASKS", "INDICATIONS_RF", "RESPONSE", "INDICATIONS_VO", "INDICATIONS_RF_VO", "GOSPROGRAM_VO",
               "TARGET_GOSPROGRAM_VO", "GOSPROGRAM_TASKS", "GOSPROGRAM_TARGET", "INDICATIONS_RF_TARGET"]


def db_conn():
    conn = psycopg2.connect(dbname='project-cs', user='cs-main',
                            password='eid4Uepo', host='cs-vm-postgre.cs.vsu.ru')
    cursor = conn.cursor()

    return cursor, conn


def fill_db(cur, conn, sheet):
    take_data_from_excel(sheet, cur, conn)

    print()
    # conn.commit()


def take_data_from_excel(sheet, cur, conn):
    # list = []
    # list_index = []
    # for i in range(6, sheet.max_row):
    #     al = sheet.cell(row=i, column=3).value
    #     index = sheet.cell(row=i, column=2).value
    #     if (al != None):
    #         list_index.append(index)
    #         list.append(al)
    #
    # cyr_list = sorted(set(list), key=list.index)
    # set_index_list = sorted(set(list_index), key=list_index.index)
    # list.clear()
    # count = 0
    # for i in range(0, len(cyr_list)):
    #     print(count)
    #
    #     # cur.execute(f"INSERT INTO CUR VALUES ({set_index_list[count]}, '{cyr_list[i]}')")
    #     count += 1
    # conn.commit()

    # ---cyr----------------------------------------------------

    # count = 1
    # for i in range(6, sheet.max_row):
    #     al = sheet.cell(row=i, column=4).value
    #     index = sheet.cell(row=i, column=2).value
    #     if (al != None):
    #         cur.execute(f"INSERT INTO TASKS VALUES ({count}, {index}, '{al}')")
    #         count += 1
    # conn.commit()

    # ---tasks-------------------------------------------------
    # count = 1
    # for i in range(6, sheet.max_row):
    #     al = sheet.cell(row=i, column=5).value
    #     index = sheet.cell(row=i, column=2).value
    #     if al != None:
    #         cur.execute(f"INSERT INTO INDICATIONS_RF VALUES ({count}, {count}, '{al}')")
    #         count += 1
    # conn.commit()

    # ---INDICATIONS_RF-----------------------------------------

    count = 1
    list = []
    for i in range(6, sheet.max_row):
        al = sheet.cell(row=i, column=8).value
        if (al != None):
            list.append(al)
    indicators_list = sorted(set(list), key=list.index)

    for i in range(0, len(indicators_list)):
        cur.execute(f"INSERT INTO RESPONSE_OBJ VALUES ({count}, '{indicators_list[i]}')")
        count += 1
    conn.commit()

    # ---RESPONSE_OBJ------------------------------------------------

    # count = 1
    # for i in range(6, sheet.max_row):
    #     al = sheet.cell(row=i, column=7).value
    #     response = sheet.cell(row=i, column=8).value
    #     cur.execute(f"SELECT id FROM RESPONSE_OBJ WHERE response_obj = '{response}'")
    #     val = cur.fetchone()
    #     if al != None:
    #         cur.execute(f"INSERT INTO INDICATIONS_VO VALUES ({count}, {val[0]}, '{al}')")
    #         count += 1
    # conn.commit()

    # ---INDICATIONS_VO--------------------------------------------

    # count = 1
    # for i in range(6, sheet.max_row):
    #     al = sheet.cell(row=i, column=6).value
    #     val = sheet.cell(row=i, column=8).value
    #     cur.execute(f"SELECT id FROM RESPONSE WHERE response_obj = '{val}'")
    #     response = cur.fetchone()
    #     val = sheet.cell(row=i, column=5).value
    #     cur.execute(f"SELECT id FROM INDICATIONS_RF WHERE ind_title_rf = '{val}'")
    #     id_ind_rf = cur.fetchone()
    #     val = sheet.cell(row=i, column=7).value
    #     cur.execute(f"SELECT id FROM INDICATIONS_VO WHERE ind_title_vo = '{val}'")
    #     ind_title_vo = cur.fetchone()
    #
    #     if al != None:
    #         cur.execute(
    #             f"INSERT INTO GOSPROGRAM_VO VALUES ({count}, '{al}', {response[0]}, {id_ind_rf[0]}, {ind_title_vo[0]})")
    #         count += 1
    # conn.commit()

    # ---GOSPROGRAM_VO---------------------------------------------


if __name__ == "__main__":
    cur, conn = db_conn()
    path = "ЦУР и ГП ВО_показатели.xlsx"
    wb = load_workbook(path)
    sheet = wb.get_sheet_by_name('Нац.набор')
    fill_db(cur, conn, sheet)
